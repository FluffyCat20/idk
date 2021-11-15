import subprocess
import json
from openpyxl import Workbook
from openpyxl.formula.translate import Translator

wb = Workbook()
error_norms_output = {}
grid_size = 100
sizes = []
sizes_number = 4

for i in range (sizes_number):
	print (grid_size, "grid is processing")
	sizes.append(str(grid_size))

	#change grid size in config
	config_file = open("config.json", 'r')
	config = json.load(config_file)
	config_file.close()
	config["nodes_number"] = grid_size
	config_file = open("config.json", 'w')
	json.dump(config, config_file, indent = 2)
	config_file.close()

	#run
	subprocess.run(["D:\\gas_dyn\\Lax_Friedrichs_method\\build_release\\lax_friedrichs.exe"])
	'''subprocess.run forces exe to create files in the folder where the py script is
			not in the folder of exe file
				surprise mf'''

	#save error norms
	error_norms = open("error_norms.json", 'r')
	error_norms_output[grid_size] = json.load(error_norms)
	error_norms.close()

	grid_size *= 10

with open ("results\\error_norms.json", 'w') as write_file:
	json.dump (error_norms_output, write_file, indent = 2)

ws_names = list(error_norms_output[100].keys())
ws = {}
for name in ws_names:
	ws[name] = wb.create_sheet(name)

for sheet in ws.values():
	sheet.append(["", ""] + sizes)

norms = list(error_norms_output[100]["whole_grid"].keys())
values = list(error_norms_output[100]["whole_grid"]["uniform"].keys())

for ws_name in ws.keys():	
	for norm in norms:
		for value in values:
			er_n = [error_norms_output[gr_s][ws_name][norm][value] for gr_s in error_norms_output.keys()]
			ws[ws_name].append([norm, value] + er_n)
			ws[ws_name].append(["decreasing rate"])
			cur_row = ws[ws_name].max_row
			first_cell =  'D' + str(cur_row)
			numerator_cell = 'C' + str(cur_row - 1)
			denominator_cell = 'D' + str(cur_row - 1) 
			ws[ws_name][first_cell] = "=" + numerator_cell + '/' + denominator_cell
			for column in range (ord('E'), ord('C') + sizes_number):
				cur_cell = chr(column) + str(cur_row)
				ws[ws_name][cur_cell] = Translator("=" + numerator_cell + '/' + denominator_cell, origin = first_cell).translate_formula(cur_cell)


wb.save("results\\error_norms_auto.xlsx")